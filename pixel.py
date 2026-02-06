import numpy as np
from PIL import Image
import pandas as pd
from collections import Counter
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def procesuj_obraz_na_excel(sciezka_pliku, max_kolorow=16):
    if not os.path.exists(sciezka_pliku):
        print(f"Błąd: Nie znaleziono pliku {sciezka_pliku}")
        return

    print("--- ROZPOCZĘCIE PRZETWARZANIA ---")
    
    # 1. Wczytanie i kwantyzacja (redukcja kolorów)
    img = Image.open(sciezka_pliku).convert('RGB')
    img_quantized = img.quantize(colors=max_kolorow, method=Image.MAXCOVERAGE).convert('RGB')
    
    # 2. Pobranie danych o pikselach i indeksowanie
    piksele = list(img_quantized.getdata())
    szerokosc, wysokosc = img_quantized.size
    
    # Sortujemy kolory od najczęstszego (będą miały indeksy od 0 w dół)
    licznik_kolorow = Counter(piksele)
    unikalne_kolory = [kolor for kolor, count in licznik_kolorow.most_common()]
    
    kolor_na_id = {kolor: i for i, kolor in enumerate(unikalne_kolory)}
    id_na_kolor = {i: kolor for kolor, i in kolor_na_id.items()}
    
    # 3. Przygotowanie macierzy NumPy
    macierz = np.zeros((wysokosc, szerokosc), dtype=int)
    piksele_array = np.array(piksele).reshape((wysokosc, szerokosc, 3))
    
    for y in range(wysokosc):
        for x in range(szerokosc):
            kolor_rgb = tuple(piksele_array[y, x])
            macierz[y, x] = kolor_na_id[kolor_rgb]

    # --- ZAPIS PLIKÓW ---

    # A. Zapis do CSV
    pd.DataFrame(macierz).to_csv("macierz_wynikowa.csv", index=False, header=False)
    print("1. Zapisano macierz do: macierz_wynikowa.csv")

    # B. Zapis Legendy (TXT)
    with open("legenda_kolorow.txt", "w") as f:
        f.write("ID; R; G; B; Ilosc_Pikseli\n")
        for i, kolor in id_na_kolor.items():
            f.write(f"{i}; {kolor[0]}; {kolor[1]}; {kolor[2]}; {licznik_kolorow[kolor]}\n")
    print("2. Zapisano legendę do: legenda_kolorow.txt")

    # C. Generowanie kolorowego Excela
    print("3. Generowanie kolorowego pliku Excel (to może chwilę potrwać)...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa Kolorow"

    for y in range(wysokosc):
        for x in range(szerokosc):
            indeks = macierz[y, x]
            kolor_rgb = id_na_kolor[indeks]
            
            # Konwersja RGB na HEX dla Excela
            hex_color = '{:02X}{:02X}{:02X}'.format(*kolor_rgb)
            
            cell = ws.cell(row=y + 1, column=x + 1)
            cell.value = indeks
            
            # Kolorowanie komórki
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.fill = fill

    # Formatowanie Excela: ustawiamy kolumny tak, by komórki były bardziej kwadratowe
    for col in range(1, szerokosc + 1):
        ws.column_dimensions[get_column_letter(col)].width = 3.5

    wb.save("podglad_kolorowy.xlsx")
    print("Sukces! Plik podglad_kolorowy.xlsx jest gotowy.")

# --- MIEJSCE NA TWOJĄ NAZWĘ PLIKU ---
# Upewnij się, że ten plik jest w tym samym folderze co skrypt!
moj_plik = 'final-image.jpg' 

if __name__ == "__main__":
    procesuj_obraz_na_excel(moj_plik)